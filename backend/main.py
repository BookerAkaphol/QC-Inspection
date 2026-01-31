import os
import sys
import io
import base64
import logging
import asyncio
import uvicorn
from datetime import datetime
from typing import Tuple, Optional, List, Dict

import cv2
import numpy as np
import pandas as pd
from fastapi import FastAPI, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from ultralytics import YOLO
from PIL import Image as PILImage
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment

# --- Logging Configuration ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# --- Application Configuration ---
MODEL_CONFIG = {
    "688-1": 21,
    "688-2": 25,
    "688-3": 22 
}
MODEL_PATH: str = "last1280.pt"
EXCEL_FILE: str = "qc_log.xlsx"
IMAGE_FOLDER: str = "captured_images"

os.makedirs(IMAGE_FOLDER, exist_ok=True)

app = FastAPI(title="QC Inspection API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/images", StaticFiles(directory=IMAGE_FOLDER), name="images")

# --- Model Initialization ---
model: Optional[YOLO] = None
if os.path.exists(MODEL_PATH):
    try:
        model = YOLO(MODEL_PATH)
        logger.info(f"Successfully loaded YOLO model from: {MODEL_PATH}")
    except Exception as e:
        logger.error(f"Failed to load model: {e}")
else:
    logger.warning("Model file not found. System running in MOCKUP mode.")


# --- Core Processing Logic ---
def process_and_annotate(img_cv: np.ndarray, model_name: str) -> Tuple[np.ndarray, int, str, int]:
    hole_count = 0
    annotated_img = img_cv.copy()
    
    target_holes = MODEL_CONFIG.get(model_name, 25)

    if model:
        results = model(annotated_img, conf=0.5, iou=0.5, agnostic_nms=True)
        for r in results:
            hole_count = len(r.boxes)
            annotated_img = r.plot(line_width=2, font_size=0.5, labels=False)
    else:
        # Mockup Mode
        hole_count = target_holes
        h, w, _ = annotated_img.shape
        cv2.rectangle(annotated_img, (50, 50), (w - 50, h - 50), (0, 255, 0), 2)
        cv2.putText(annotated_img, "MOCKUP MODE", (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

    # Determine Quality Status
    status = "PASS" if hole_count == target_holes else "NG"

    # Overlay status text
    text_color = (0, 255, 0) if status == "PASS" else (0, 0, 255)
    cv2.putText(
        annotated_img,
        f"{status} ({hole_count}/{target_holes})",
        (20, 40),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.7,
        text_color,
        2
    )

    return annotated_img, hole_count, status, target_holes


def encode_image_to_base64(cv_img: np.ndarray) -> str:
    _, buffer = cv2.imencode('.jpg', cv_img)
    return base64.b64encode(buffer).decode('utf-8')


def save_record(image_bytes: bytes, count: int, status: str, model_name: str) -> None:
    try:
        now = datetime.now()
        timestamp_str = now.strftime('%Y%m%d_%H%M%S')
        display_dt = now.strftime("%Y-%m-%d %H:%M:%S")
        filename = f"qc_{timestamp_str}.jpg"
        filepath = os.path.join(IMAGE_FOLDER, filename)

        with open(filepath, "wb") as f:
            f.write(image_bytes)

        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Datetime", "Model", "Status", "Count", "Image", "Filename"])

        next_row = ws.max_row + 1
        ws.row_dimensions[next_row].height = 80
        center_align = Alignment(horizontal='center', vertical='center')

        ws.cell(row=next_row, column=1, value=display_dt).alignment = center_align
        ws.cell(row=next_row, column=2, value=model_name).alignment = center_align
        ws.cell(row=next_row, column=3, value=status).alignment = center_align
        ws.cell(row=next_row, column=4, value=count).alignment = center_align
        ws.cell(row=next_row, column=6, value=filename).alignment = center_align

        try:
            img_stream = io.BytesIO(image_bytes)
            pil_img = PILImage.open(img_stream)
            pil_img.thumbnail((140, 100))
            img_byte_arr = io.BytesIO()
            pil_img.save(img_byte_arr, format='JPEG')
            excel_img = ExcelImage(img_byte_arr)
            excel_img.anchor = f"E{next_row}"
            ws.add_image(excel_img)
        except Exception as e:
            logger.error(f"Failed to embed image: {e}")

        wb.save(EXCEL_FILE)
        logger.info(f"Record saved: {filename}")

    except Exception as e:
        logger.error(f"Critical error save: {e}")


# --- API Endpoints ---
@app.post("/preview", tags=["Inspection"])
async def preview_inspect(
    file: UploadFile = File(...), 
    model_name: str = Form(...)
):
    try:
        contents = await file.read()
        nparr = np.frombuffer(contents, np.uint8)
        img_cv = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        annotated_img, count, status, target = process_and_annotate(img_cv, model_name)
        base64_str = encode_image_to_base64(annotated_img)

        return JSONResponse({
            "status": status,
            "count": count,
            "required": target,
            "image_base64": base64_str
        })
    except Exception as e:
        logger.error(f"Error preview: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/save", tags=["Inspection"])
async def save_result(
    file: UploadFile = File(...), 
    model_name: str = Form(...)
):
    try:
        contents = await file.read()
        nparr = np.frombuffer(contents, np.uint8)
        img_cv = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        annotated_img, count, status, _ = process_and_annotate(img_cv, model_name)
        
        _, img_encoded = cv2.imencode('.jpg', annotated_img)
        final_image_bytes = img_encoded.tobytes()

        save_record(final_image_bytes, count, status, model_name)
        return {"message": "Saved successfully."}
    except Exception as e:
        logger.error(f"Error save: {e}")
        return JSONResponse(status_code=500, content={"error": "Failed to save."})


@app.get("/logs", tags=["Data Access"])
def get_logs():
    if not os.path.exists(EXCEL_FILE): return []
    try:
        df = pd.read_excel(EXCEL_FILE)
        target_cols = ['Datetime', 'Model', 'Status', 'Count', 'Filename']
        existing_cols = [c for c in target_cols if c in df.columns]
        data = df[existing_cols].tail(50).iloc[::-1].fillna('')
        return data.to_dict(orient="records")
    except Exception as e:
        logger.error(f"Error logs: {e}")
        return []


@app.get("/download-excel", tags=["Data Access"])
def download_excel():
    if os.path.exists(EXCEL_FILE):
        return FileResponse(path=EXCEL_FILE, filename="QC_Report.xlsx", media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return JSONResponse(status_code=404, content={"error": "Not found."})


if os.path.exists("dist"):
    app.mount("/", StaticFiles(directory="dist", html=True), name="static")

if __name__ == "__main__":
    if sys.platform.startswith("win"):
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    ssl_config = {}
    if os.path.exists("key.pem") and os.path.exists("cert.pem"):
        logger.info("Starting HTTPS...")
        ssl_config = {"ssl_keyfile": "key.pem", "ssl_certfile": "cert.pem"}
    
    uvicorn.run(app, host="0.0.0.0", port=8000, **ssl_config)